# Flask Requirements
from flask import Flask, request, render_template, jsonify, Response, redirect, url_for, flash, session
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from urllib.parse import urlparse, urljoin
# Threading and Request
import threading
import requests
# Database System
import sqlite3
# Timezone Requirements
import datetime
import time
from datetime import timedelta
from zoneinfo import ZoneInfo  # Untuk timezone Indonesia
# MQTT Service 
import paho.mqtt.client as mqtt
# Telegram Requirements
from telegram import InlineKeyboardButton, InlineKeyboardMarkup, Update, Bot
from telegram.ext import ContextTypes, ApplicationBuilder, CallbackQueryHandler, MessageHandler, filters
import asyncio
# Operatin System Requirements
from dotenv import load_dotenv
import os
# Logging System
import logging
# Excel Requirements
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
# Untuk Input dan Output
from queue import Queue, Empty
from io import BytesIO
from functools import wraps
import json

# Setup logging dengan timezone Indonesia
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

load_dotenv()

# --- LOGIN SYSTEM: User Class ---
class User(UserMixin):
    def __init__(self, id, username, password):
        self.id = id
        self.username = username
        self.password = password

def is_safe_url(target):
    """Validasi URL untuk mencegah open redirect vulnerability"""
    ref_url = urlparse(request.host_url)
    test_url = urlparse(urljoin(request.host_url, target))
    return test_url.scheme in ('http', 'https') and ref_url.netloc == test_url.netloc

# 2. SESSION TIMEOUT DECORATOR (tambah sebelum class TemperatureMonitor)
def check_session_timeout(f):
    """Decorator untuk mengecek apakah session sudah timeout"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if current_user.is_authenticated:
            # Cek apakah ada timestamp login di session
            if 'login_timestamp' in session:
                login_time = session['login_timestamp']
                current_time = time.time()
                
                # Hitung durasi login (24 jam = 86400 detik)
                session_duration = current_time - login_time
                max_session_duration = 24 * 60 * 60  # 24 jam dalam detik
                
                if session_duration > max_session_duration:
                    # Session expired, logout otomatis
                    logout_user()
                    session.clear()  # Bersihkan semua session data
                    flash('Your session has expired after 24 hours. Please log in again.', 'warning')
                    return redirect(url_for('login'))
                else:
                    # Session masih valid, update last activity
                    session['last_activity'] = current_time
            else:
                # Tidak ada timestamp login, anggap session tidak valid
                logout_user()
                session.clear()
                flash('Invalid session. Please log in again.', 'warning')
                return redirect(url_for('login'))
        
        return f(*args, **kwargs)
    return decorated_function

# 3. SESSION INFO HELPER (tambah sebelum class TemperatureMonitor)
def get_session_info():
    """Helper untuk mendapatkan informasi session"""
    if not current_user.is_authenticated or 'login_timestamp' not in session:
        return None
    
    login_time = session['login_timestamp']
    current_time = time.time()
    session_age = current_time - login_time
    remaining_time = (24 * 60 * 60) - session_age  # sisa waktu dalam detik
    
    return {
        'login_time': datetime.datetime.fromtimestamp(login_time).strftime('%Y-%m-%d %H:%M:%S'),
        'session_age_hours': session_age / 3600,
        'remaining_hours': max(0, remaining_time / 3600),
        'remaining_minutes': max(0, (remaining_time % 3600) / 60),
        'is_expiring_soon': remaining_time < (2 * 3600),  # kurang dari 2 jam
        'expires_at': datetime.datetime.fromtimestamp(login_time + (24 * 60 * 60)).strftime('%Y-%m-%d %H:%M:%S')
    }
# ------------------------------

# Basic Configuration
class TemperatureMonitorConfig:
    """Class untuk mengelola konfigurasi aplikasi"""
    
    def __init__(self):
        self.MQTT_BROKER = os.getenv("MQTT_BROKER")
        self.MQTT_PORT = 1883
        self.MQTT_TOPICS = {
            "dryer1": os.getenv("MQTT_TOPIC_1"),
            "dryer2": os.getenv("MQTT_TOPIC_2"),
            "dryer3": os.getenv("MQTT_TOPIC_3"),
        }
        self.TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN")
        self.CHAT_ID = os.getenv("CHAT_ID")
        self.DATA_SAVE_INTERVAL = 600 # Pengambilan data di lakukan setiap 10 menit sekali
        self.TEMPERATURE_OFFSET = 12.6
        self.INDONESIA_TZ = ZoneInfo("Asia/Jakarta")
        self.MIN_TEMP_ALERT = float(120)
        self.MAX_TEMP_ALERT = float(155)
        self.DB_PATH = "/data/data_suhu_multi.db" if os.path.exists("/data") else "data_suhu_multi.db"
        
        self.validate()
        
    def validate(self):
        """Validasi konfigurasi yang diperlukan"""
        if not all([self.MQTT_BROKER, self.TELEGRAM_TOKEN, self.CHAT_ID]):
            logger.error("Missing required environment variables")
            exit(1)
            
        logger.info(f"Config loaded - Broker: {self.MQTT_BROKER}")

    def get_indonesia_time(self):
        """Get current time in Indonesia timezone"""
        return datetime.datetime.now(self.INDONESIA_TZ)

    def format_indonesia_time(self, dt=None):
        """Format time in Indonesian format with timezone"""
        if dt is None:
            dt = self.get_indonesia_time()
        return dt.strftime("%Y-%m-%d %H:%M:%S %Z")

    def format_indonesia_time_simple(self, dt=None):
        """Format time in simple format without timezone for database"""
        if dt is None:
            dt = self.get_indonesia_time()
        return dt.strftime("%Y-%m-%d %H:%M:%S")

    def apply_temperature_offset(self, raw_temp):
        """Apply consistent temperature offset"""
        if raw_temp is None:
            return None
        return raw_temp + self.TEMPERATURE_OFFSET

class DatabaseManager:
    """Class untuk mengelola operasi database untuk multi-dryer"""
    
    def __init__(self, db_path):
        self.db_path = db_path
        self.initialize_database()
        
    def initialize_database(self):
        """Initialize database dengan table yang bisa menyimpan ID dryer"""
        with self.get_connection() as conn:
            c = conn.cursor()
            c.execute("""
                CREATE TABLE IF NOT EXISTS suhu (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    waktu TEXT,
                    dryer_id TEXT, 
                    suhu REAL
                )
            """)
            # --- LOGIN SYSTEM: Create users table ---
            c.execute("""
                CREATE TABLE IF NOT EXISTS users (
                    id INTEGER PRIMARY KEY AUTOINCREMENT, username TEXT UNIQUE NOT NULL, password TEXT NOT NULL
                )
            """)
            # ----------------------------------------
            
        logger.info(f"Database multi-dryer initialized at: {self.db_path}")
    
    def get_connection(self):
        """Mendapatkan koneksi database yang thread-safe"""
        conn = sqlite3.connect(self.db_path, check_same_thread=False, timeout=30.0)
        conn.execute("PRAGMA journal_mode=WAL")
        return conn
    
    # --- LOGIN SYSTEM: Methods for user management ---
    def get_user_by_username(self, username):
        try:
            with self.get_connection() as conn:
                c = conn.cursor()
                c.execute("SELECT * FROM users WHERE username = ?", (username,))
                user_data = c.fetchone()
                if user_data:
                    return User(id=user_data[0], username=user_data[1], password=user_data[2])
            return None
        except Exception as e:
            logger.error(f"Error getting user by username: {e}")
            return None

    def get_user_by_id(self, user_id):
        try:
            with self.get_connection() as conn:
                c = conn.cursor()
                c.execute("SELECT * FROM users WHERE id = ?", (user_id,))
                user_data = c.fetchone()
                if user_data:
                    return User(id=user_data[0], username=user_data[1], password=user_data[2])
            return None
        except Exception as e:
            logger.error(f"Error getting user by ID: {e}")
            return None
    
    def create_initial_user(self, username, password):
        """Membuat user awal jika belum ada, menggunakan Fly secrets."""
        if not self.get_user_by_username(username):
            logger.info(f"User '{username}' tidak ditemukan, mencoba membuat user baru...")
            try:
                with self.get_connection() as conn:
                    c = conn.cursor()
                    hashed_password = generate_password_hash(password, method='pbkdf2:sha256')
                    c.execute("INSERT INTO users (username, password) VALUES (?, ?)", (username, hashed_password))
                    conn.commit()
                    logger.info(f"User '{username}' berhasil dibuat dari environment secrets.")
            except Exception as e:
                logger.error(f"Gagal membuat initial user: {e}")
    # -----------------------------------------------
    
    def insert_temperature(self, waktu, dryer_id, suhu):
        """Insert data suhu ke database dengan menyertakan ID dryer"""
        try:
            with self.get_connection() as conn:
                c = conn.cursor()
                c.execute("INSERT INTO suhu (waktu, dryer_id, suhu) VALUES (?, ?, ?)", (waktu, dryer_id, suhu))
            return True
        except Exception as e:
            logger.error(f"Error inserting temperature: {e}")
            return False
    
    def get_data_by_date_pivoted(self, date_str, latest_only=False):
        """
        Mendapatkan data untuk tanggal tertentu.
        Jika latest_only=True, hanya mengembalikan 1 baris data terbaru.
        """
        try:
            start_time = f"{date_str} 00:00:00"
            end_time = f"{date_str} 23:59:59"
            
            sql = """
            SELECT
                strftime('%Y-%m-%d %H:%M:%S', waktu) as timestamp,
                MAX(CASE WHEN dryer_id = 'dryer1' THEN suhu END) as dryer1_suhu,
                MAX(CASE WHEN dryer_id = 'dryer2' THEN suhu END) as dryer2_suhu,
                MAX(CASE WHEN dryer_id = 'dryer3' THEN suhu END) as dryer3_suhu
            FROM suhu
            WHERE waktu BETWEEN ? AND ?
            GROUP BY timestamp
            """
            
            if latest_only:
                sql += " ORDER BY timestamp DESC LIMIT 1"
            else:
                sql += " ORDER BY timestamp ASC"

            with self.get_connection() as conn:
                c = conn.cursor()
                c.execute(sql, (start_time, end_time))
                return c.fetchall()
        except Exception as e:
            logger.error(f"Error getting pivoted data for date {date_str}: {e}")
            return []
            
    def get_data_since(self, since_time):
        """Mendapatkan data sejak waktu tertentu"""
        try:
            with self.get_connection() as conn:
                c = conn.cursor()
                c.execute("SELECT * FROM suhu WHERE datetime(waktu) >= datetime(?) ORDER BY waktu", (since_time,))
                return c.fetchall()
        except Exception as e:
            logger.error(f"Error getting data since {since_time}: {e}")
            return []
    
    def get_recent_data(self, limit=5):
        """Mendapatkan data terbaru (untuk handler Telegram)"""
        try:
            with self.get_connection() as conn:
                c = conn.cursor()
                c.execute("SELECT waktu, dryer_id, suhu FROM suhu ORDER BY id DESC LIMIT ?", (limit,))
                return c.fetchall()
        except Exception as e:
            logger.error(f"Error getting recent data: {e}")
            return []

class MQTTService:
    """Class untuk mengelola koneksi dan komunikasi MQTT"""
    
    def __init__(self, config, data_callback):
        self.config = config
        self.data_callback = data_callback
        self.client = mqtt.Client()
        self.is_connected = False
        self.setup_callbacks()
        
    def setup_callbacks(self):
        """Setup callback functions untuk MQTT client"""
        self.client.on_connect = self._on_connect
        self.client.on_message = self._on_message
        self.client.on_disconnect = self._on_disconnect
    
    def _on_connect(self, client, userdata, flags, rc):
        """Callback ketika terhubung ke MQTT broker"""
        if rc == 0:
            self.is_connected = True
            logger.info(f"MQTT Connected successfully at {self.config.format_indonesia_time()}")
            for topic in self.config.MQTT_TOPICS.values():
                if topic: self.client.subscribe(topic)
            logger.info(f"Subscribed to topics: {self.config.MQTT_TOPICS}")
        else:
            logger.error(f"MQTT Connection failed with code {rc}")
    
    def _on_message(self, client, userdata, msg):
        """Callback ketika menerima message MQTT"""
        try:
            raw_suhu = float(msg.payload.decode())
            if self.data_callback:
                self.data_callback(raw_suhu, msg.topic)
        except Exception as e:
            logger.error(f"Error parsing MQTT data: {e}")
    
    def _on_disconnect(self, client, userdata, rc):
        """Callback ketika terputus dari MQTT broker"""
        self.is_connected = False
        logger.warning(f"MQTT Disconnected with code {rc} at {self.config.format_indonesia_time()}")
    
    def connect(self):
        """Connect ke MQTT broker"""
        try:
            self.client.connect(self.config.MQTT_BROKER, self.config.MQTT_PORT, 60)
            self.client.loop_start()
            logger.info("MQTT Client started")
            return True
        except Exception as e:
            logger.error(f"MQTT Connection failed: {e}")
            return False
    
    def disconnect(self):
        """Disconnect dari MQTT broker"""
        self.client.loop_stop()
        self.client.disconnect()

class TelegramService:
    """Class untuk mengelola komunikasi Telegram dengan antrean dan worker."""
    def __init__(self, config, db_manager):
        self.config = config
        self.db_manager = db_manager
        self.bot = Bot(token=config.TELEGRAM_TOKEN)
        self.application = ApplicationBuilder().token(self.config.TELEGRAM_TOKEN).build()
        self._setup_handlers()
        self.message_queue = Queue()
        self.worker_thread = None
        self.is_worker_running = False

    def _setup_handlers(self):
        self.application.add_handler(MessageHandler(filters.Regex('^Mulai$'), self.start))
        self.application.add_handler(CallbackQueryHandler(self.button))

    async def _send_message_async(self, message):
        try:
            await self.bot.send_message(chat_id=self.config.CHAT_ID, text=message, parse_mode="Markdown")
            logger.info("Pesan Telegram berhasil dikirim dari worker.")
        except Exception as e:
            logger.error(f"Gagal mengirim pesan dari worker: {e}")

    async def _send_document_async(self, file_path, caption):
        try:
            with open(file_path, "rb") as file:
                await self.bot.send_document(chat_id=self.config.CHAT_ID, document=file, caption=caption, parse_mode="Markdown")
            logger.info(f"Dokumen {file_path} berhasil dikirim dari worker.")
        except Exception as e:
            logger.error(f"Gagal mengirim dokumen dari worker: {e}")

    def _process_queue(self):
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        while self.is_worker_running:
            try:
                task_type, *args = self.message_queue.get(timeout=1) 
                if task_type == 'message':
                    loop.run_until_complete(self._send_message_async(args[0]))
                elif task_type == 'document':
                    loop.run_until_complete(self._send_document_async(args[0], args[1]))
                self.message_queue.task_done()
            except Empty:
                continue
            except Exception as e:
                logger.error(f"Error di dalam Telegram worker thread: {e}")

    def start_worker(self):
        if not self.is_worker_running:
            self.is_worker_running = True
            self.worker_thread = threading.Thread(target=self._process_queue, daemon=True, name="TelegramWorker")
            self.worker_thread.start()
            logger.info("Telegram worker thread dimulai.")

    def stop_worker(self):
        self.is_worker_running = False
        if self.worker_thread:
            self.worker_thread.join(timeout=5)
            logger.info("Telegram worker thread dihentikan.")

    def send_message(self, message):
        self.message_queue.put(('message', message))

    def send_document(self, file_path, caption):
        self.message_queue.put(('document', file_path, caption))
    
    async def start(self, update, context):
        keyboard = [
            [InlineKeyboardButton("Test Message", callback_data="test")],
            [InlineKeyboardButton("Data Dryers", callback_data="data")],
            [InlineKeyboardButton("Force Excel", callback_data="force_excel")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await update.message.reply_text("Pilih opsi:", reply_markup=reply_markup)
    
    async def button(self, update, context):
        query = update.callback_query
        await query.answer()
        if query.data == "test":
            await self._handle_test(query)
        elif query.data == "data":
            await self._handle_data(query)
        elif query.data == "force_excel":
            await self._handle_force_excel(query)
    
    async def _handle_test(self, query):
        current_time = self.config.format_indonesia_time()
        test_message = f"🧪 Test Message\n🕐 Waktu: {current_time}\n✅ Bot berfungsi normal!"
        await query.edit_message_text(test_message)
    
    async def _handle_data(self, query):
        recent_data = self.db_manager.get_recent_data(5)
        if recent_data:
            data_text = "📊 5 Data Terakhir:\n\n"
            for row in recent_data:
                data_text += f"*{row[1].upper()}*:\n🕐 {row[0]} WIB\n🌡️ {row[2]:.1f}°C\n\n"
        else:
            data_text = "❌ Tidak ada data tersedia"
        await query.edit_message_text(data_text)
    
    async def _handle_force_excel(self, query):
        await query.edit_message_text("📊 Generating Excel... Please wait...")
        today_str = self.config.get_indonesia_time().strftime('%Y-%m-%d')
        rows = self.db_manager.get_data_by_date_pivoted(today_str)
        if not rows:
            await query.edit_message_text("❌ No data available for today.")
            return
        wb = Workbook()
        ws = wb.active
        ws.title = f"Data Suhu {today_str}"
        ws.append(["Waktu (WIB)", "Dryer 1 (°C)", "Dryer 2 (°C)", "Dryer 3 (°C)"])
        for row in rows: ws.append(list(row))
        temp_dir = "/tmp" if os.path.exists("/tmp") else "."
        filename = os.path.join(temp_dir, f"manual_report_{today_str}.xlsx")
        wb.save(filename)
        caption = f"📊 Laporan Manual - {today_str}"
        with open(filename, "rb") as file:
            await self.bot.send_document(chat_id=query.message.chat_id, document=file, caption=caption)
        try: os.remove(filename)
        except: pass
        await query.edit_message_text(f"✅ Excel sent! {len(rows)} records")
    
    def start_polling(self):
        if self.application:
            self.application.run_polling()

class BackgroundTask:
    def __init__(self, interval, name="BackgroundTask"):
        self.interval = interval
        self.name = name
        self.thread = None
        self.is_running = False
    def task(self):
        raise NotImplementedError
    def run(self):
        self.is_running = True
        while self.is_running:
            try:
                if self.is_running: self.task()
            except Exception as e:
                logger.error(f"Error in {self.name}: {e}")
            if self.interval > 0:
                time.sleep(self.interval)
    def start(self):
        self.thread = threading.Thread(target=self.run, daemon=True, name=self.name)
        self.thread.start()
    def stop(self):
        self.is_running = False
        if self.thread: self.thread.join(timeout=5)

# --- DIKEMBALIKAN: DataSaveTask untuk menyimpan per 10 menit ---
class DataSaveTask(BackgroundTask):
    """Task untuk menyimpan data ke database setiap 10 menit sekali."""
    
    def __init__(self, config, data_provider, db_manager):
        super().__init__(config.DATA_SAVE_INTERVAL, "DataSaveTask")
        self.config = config
        self.data_provider = data_provider
        self.db_manager = db_manager
    
    def task(self):
        """Menyimpan data suhu terakhir dari memori ke database."""
        latest_temps = self.data_provider.get_latest_temperatures()
        waktu = self.config.format_indonesia_time_simple()
        
        logger.info(f"Menjalankan DataSaveTask pada {waktu}. Menyimpan data terakhir...")
        
        for dryer_id, temp in latest_temps.items():
            if temp is not None:
                success = self.db_manager.insert_temperature(waktu, dryer_id, temp)
                if success:
                    logger.info(f"Data tersimpan untuk {dryer_id}: {waktu} | {temp:.2f}°C")
                else:
                    logger.error(f"Gagal menyimpan data untuk {dryer_id}")
# --------------------------------------------------------------------

class DailyExcelReportTask(BackgroundTask):
    def __init__(self, config, db_manager, telegram_service):
        super().__init__(-1, "DailyExcelReportTask")
        self.config = config
        self.db_manager = db_manager
        self.telegram_service = telegram_service
    def run(self):
        self.is_running = True
        logger.info(f"Starting {self.name}. Laporan akan dikirim setiap pukul 00:00.")
        while self.is_running:
            try:
                now = self.config.get_indonesia_time()
                tomorrow = now + datetime.timedelta(days=1)
                midnight = tomorrow.replace(hour=0, minute=0, second=5, microsecond=0)
                seconds_to_wait = (midnight - now).total_seconds()
                logger.info(f"Laporan harian berikutnya dalam {seconds_to_wait / 3600:.2f} jam.")
                sleep_end_time = time.time() + seconds_to_wait
                while time.time() < sleep_end_time and self.is_running:
                    time.sleep(1)
                if self.is_running: self.task()
            except Exception as e:
                logger.error(f"Error di dalam loop {self.name}: {e}")
                time.sleep(300)
    def task(self):
        yesterday = (self.config.get_indonesia_time() - datetime.timedelta(days=1))
        yesterday_str = yesterday.strftime('%Y-%m-%d')
        logger.info(f"Memulai pembuatan laporan Excel untuk tanggal: {yesterday_str}")
        rows = self.db_manager.get_data_by_date_pivoted(yesterday_str)
        if not rows:
            logger.warning(f"Tidak ada data untuk dilaporkan pada tanggal {yesterday_str}.")
            return
        wb = Workbook()
        ws = wb.active
        ws.title = f"Data Suhu {yesterday_str}"
        ws.append(["Waktu (WIB)", "Dryer 1 (°C)", "Dryer 2 (°C)", "Dryer 3 (°C)"])
        for row in rows: ws.append(list(row))
        temp_dir = "/tmp" if os.path.exists("/tmp") else "."
        filename = os.path.join(temp_dir, f"laporan_harian_{yesterday_str}.xlsx")
        wb.save(filename)
        caption = f"📊 *Laporan Harian Suhu - {yesterday.strftime('%d %B %Y')}*"
        self.telegram_service.send_document(filename, caption)
        time.sleep(10)
        try: os.remove(filename)
        except OSError: pass

class KeepaliveTask(BackgroundTask):
    def __init__(self, config):
        super().__init__(1800, "KeepaliveTask")
        self.config = config
    def task(self):
        app_url = os.getenv("FLY_APP_NAME", "")
        if app_url:
            try:
                requests.get(f"https://{app_url}.fly.dev/keepalive", timeout=10)
            except Exception as e:
                logger.error(f"Keepalive error: {e}")

class MonitorDataTask(BackgroundTask):
    def __init__(self, config, db_manager, telegram_service):
        super().__init__(3600, "MonitorDataTask")
        self.config = config
        self.db_manager = db_manager
        self.telegram_service = telegram_service
        self.is_error_notified = False
    def task(self):
        waktu_awal = (self.config.get_indonesia_time() - datetime.timedelta(hours=1)).strftime("%Y-%m-%d %H:%M:%S")
        rows = self.db_manager.get_data_since(waktu_awal)
        if not rows: return
        unique_values = set([round(row[3], 2) for row in rows if len(row) > 3 and row[3] is not None])
        if len(unique_values) == 1:
            if not self.is_error_notified:
                suhu_error = list(unique_values)[0]
                error_message = f"⚠️ *PERINGATAN SISTEM ERROR* ⚠️\n\nSuhu macet di *{suhu_error:.2f}°C*."
                self.telegram_service.send_message(error_message)
                self.is_error_notified = True
        else:
            if self.is_error_notified:
                self.is_error_notified = False

class TemperatureMonitor:
    def __init__(self):
        self.config = TemperatureMonitorConfig()
        self.latest_temperatures = { "dryer1": None, "dryer2": None, "dryer3": None }
        self.data_lock = threading.Lock()
        self.alert_status = { "dryer1": "NORMAL", "dryer2": "NORMAL", "dryer3": "NORMAL" }
        self.db_manager = DatabaseManager(self.config.DB_PATH)
        self.telegram_service = TelegramService(self.config, self.db_manager)
        self.mqtt_service = MQTTService(self.config, self._on_mqtt_message)
        self.tasks = []
    
    # --- DIPERBAIKI: _on_mqtt_message HANYA MENGUPDATE MEMORI ---
    def _on_mqtt_message(self, raw_temperature, topic):
        """Callback untuk setiap pesan MQTT. Memproses suhu dan mengirim notifikasi."""
        dryer_id = None
        for id, t in self.config.MQTT_TOPICS.items():
            if t == topic:
                dryer_id = id
                break

        if not dryer_id:
            return

        adjusted_temperature = self.config.apply_temperature_offset(raw_temperature)

        with self.data_lock:
            self.latest_temperatures[dryer_id] = adjusted_temperature

        logger.info(f"Data {{{dryer_id}}} diterima: {adjusted_temperature:.2f}°C")

        waktu_kejadian = self.config.format_indonesia_time()
        telegram_message = None
        notification_payload = None  # Payload untuk notifikasi web

        # Cek suhu tinggi
        if adjusted_temperature > self.config.MAX_TEMP_ALERT:
            if self.alert_status[dryer_id] != "HIGH":
                title = f"🔥 Suhu Tinggi ({dryer_id.upper()})"
                message = f"Suhu mencapai {adjusted_temperature:.1f}°C, melebihi batas normal {self.config.MAX_TEMP_ALERT}°C."

                telegram_message = f"*{title}*\n\n🌡️ Suhu: *{adjusted_temperature:.1f}°C*\n🕒 Waktu: {waktu_kejadian}"
                notification_payload = {"title": title, "message": message}
                self.alert_status[dryer_id] = 'HIGH'

        # Cek suhu rendah
        elif adjusted_temperature < self.config.MIN_TEMP_ALERT:
            if self.alert_status[dryer_id] != "LOW":
                title = f"❄️ Suhu Rendah ({dryer_id.upper()})"
                message = f"Suhu turun menjadi {adjusted_temperature:.1f}°C, di bawah batas normal {self.config.MIN_TEMP_ALERT}°C."

                telegram_message = f"*{title}*\n\n🌡️ Suhu: *{adjusted_temperature:.1f}°C*\n🕒 Waktu: {waktu_kejadian}"
                notification_payload = {"title": title, "message": message}
                self.alert_status[dryer_id] = 'LOW'

        # Suhu kembali normal
        else:
            if self.alert_status[dryer_id] != 'NORMAL':
                self.alert_status[dryer_id] = 'NORMAL'

        # Kirim notifikasi jika ada
        if telegram_message:
            current_hour = self.config.get_indonesia_time().hour
            if 6 <= current_hour < 17:
                self.telegram_service.send_message(telegram_message)

        if notification_payload:
            self.notification_queue.put(notification_payload)

    def get_latest_temperatures(self):
        """Mendapatkan semua suhu terbaru dari memori."""
        with self.data_lock:
            return self.latest_temperatures.copy()
    
    # --- DIPERBAIKI: Menjalankan kembali DataSaveTask ---
    def start_background_tasks(self):
        """Memulai semua background tasks"""
        self.tasks.append(DataSaveTask(self.config, self, self.db_manager))
        self.tasks.append(DailyExcelReportTask(self.config, self.db_manager, self.telegram_service))
        self.tasks.append(KeepaliveTask(self.config))
        self.tasks.append(MonitorDataTask(self.config, self.db_manager, self.telegram_service))
        
        for task in self.tasks:
            task.start()
            
        logger.info("All background tasks started")
    
    def stop_background_tasks(self):
        for task in self.tasks:
            task.stop()
        logger.info("All background tasks stopped")
    
    def create_flask_app(self):
        app = Flask(__name__, template_folder='templates', static_folder='static')
        app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'default-secret-key-for-dev')

        # ✅ SESSION CONFIGURATION untuk 24 jam
        app.config.update(
            PERMANENT_SESSION_LIFETIME=timedelta(hours=24),  # 24 jam
            SESSION_COOKIE_HTTPONLY=True,  # Security: tidak bisa diakses via JavaScript
            SESSION_COOKIE_SECURE=False,   # Set True jika pakai HTTPS
            SESSION_COOKIE_SAMESITE='Lax'  # CSRF protection
        )

        login_manager = LoginManager()
        login_manager.init_app(app)
        login_manager.login_view = 'login'
        login_manager.session_protection = "strong"

        # --- Login manager
        @login_manager.user_loader
        def load_user(user_id):
            return self.db_manager.get_user_by_id(user_id)
        
        #Custom unauthorized handler untuk smart redirect
        @login_manager.unauthorized_handler
        def unauthorized():
            if request.endpoint != 'login':
                return redirect(url_for('login', next=request.url))
            return redirect(url_for('login'))

        # Index.html
        @app.route("/")
        @login_required
        @check_session_timeout  
        def index():
            with self.data_lock:
                context = {
                    "current_suhu_1": f"{self.latest_temperatures['dryer1']:.1f} °C" if self.latest_temperatures['dryer1'] else "N/A",
                    "current_suhu_2": f"{self.latest_temperatures['dryer2']:.1f} °C" if self.latest_temperatures['dryer2'] else "N/A",
                    "current_suhu_3": f"{self.latest_temperatures['dryer3']:.1f} °C" if self.latest_temperatures['dryer3'] else "N/A",
                    "current_time": self.config.format_indonesia_time(),
                    "timezone": str(self.config.INDONESIA_TZ)
                }
            return render_template("index.html", **context)

        # Dwidaya
        @app.route("/dwidaya")
        @login_required
        @check_session_timeout  
        def dwidaya():
            with self.data_lock:
                context = {
                    "current_suhu_1": f"{self.latest_temperatures['dryer1']:.1f} °C" if self.latest_temperatures['dryer1'] else "N/A",
                    "current_suhu_2": f"{self.latest_temperatures['dryer2']:.1f} °C" if self.latest_temperatures['dryer2'] else "N/A",
                    "current_suhu_3": f"{self.latest_temperatures['dryer3']:.1f} °C" if self.latest_temperatures['dryer3'] else "N/A",
                    "current_time": self.config.format_indonesia_time(),
                    "timezone": str(self.config.INDONESIA_TZ)
                }
            return render_template("dwidaya.html", **context)

        @app.route('/login', methods=['GET', 'POST'])
        def login():
            if current_user.is_authenticated:
                return redirect(url_for('index'))
            
            if request.method == 'POST':
                username = request.form['username']
                password = request.form['password']
                user = self.db_manager.get_user_by_username(username)
                
                if user and check_password_hash(user.password, password):
                    # ✅ LOGIN USER dengan permanent session
                    login_user(user, remember=False)  # Tidak pakai "remember me"
                    session.permanent = True  # Set session sebagai permanent
                    
                    # ✅ SIMPAN TIMESTAMP LOGIN
                    session['login_timestamp'] = time.time()
                    session['last_activity'] = time.time()
                    session['username'] = user.username
                    
                    # Smart redirect
                    next_page = request.args.get('next')
                    if not next_page or not is_safe_url(next_page):
                        next_page = url_for('index')
                    
                    flash(f'Welcome back, {user.username}! Session valid for 24 hours.', 'success')
                    logger.info(f"User {user.username} logged in successfully from {request.remote_addr}")
                    
                    return redirect(next_page)
                else:
                    flash('Username atau password salah', 'danger')
                    logger.warning(f"Failed login attempt for username: {username} from {request.remote_addr}")
            
            return render_template('login.html')
    
        @app.route('/logout')
        @login_required
        def logout():
            logout_user()
            flash('You have been logged out successfully.', 'info')
            return redirect(url_for('login'))
        
            def generate():
                try:
                    while True:
                        try:
                            # Ambil notifikasi dari antrean dengan timeout
                            notification = self.notification_queue.get(timeout=30)  # 30 detik timeout
                            # Format sebagai Server-Sent Event (SSE)
                            yield f"data: {json.dumps(notification)}\n\n"
                            self.notification_queue.task_done()
                        except Empty:
                            # Kirim heartbeat setiap 30 detik untuk menjaga koneksi
                            yield f"data: {json.dumps({'heartbeat': True})}\n\n"
                        except Exception as e:
                            logger.error(f"Error in notification stream: {e}")
                            yield f"data: {json.dumps({'error': str(e)})}\n\n"
                            time.sleep(1)
                except Exception as e:
                    logger.error(f"Fatal error in notification stream: {e}")
                    yield f"data: {json.dumps({'error': 'Stream terminated'})}\n\n"
            
            return Response(generate(), mimetype='text/event-stream')

        # ---- Stream Data
        @app.route("/stream-data")
        @login_required
        def stream_data():
            def generate_data():
                """
                Generator function yang akan mengirim data suhu terbaru.
                """
                try:
                    while True:
                        # Ambil data suhu terbaru dari memori
                        with self.data_lock:
                            latest_temps = self.latest_temperatures.copy()

                        # Format data menjadi JSON
                        data_payload = {
                            "dryer1": f"{latest_temps['dryer1']:.1f}" if latest_temps['dryer1'] is not None else "N/A",
                            "dryer2": f"{latest_temps['dryer2']:.1f}" if latest_temps['dryer2'] is not None else "N/A",
                            "dryer3": f"{latest_temps['dryer3']:.1f}" if latest_temps['dryer3'] is not None else "N/A",
                        }
                        
                        # Kirim data dalam format SSE: "data: <json_string>\n\n"
                        yield f"data: {json.dumps(data_payload)}\n\n"
                        
                        # Tunggu sebentar sebelum mengirim data berikutnya untuk efisiensi
                        time.sleep(2) # Kirim update setiap 2 detik
                except GeneratorExit:
                    # Ini akan terjadi jika klien menutup koneksi
                    logger.info("Koneksi stream data ditutup oleh klien.")

            # Kembalikan response dengan tipe mimetype 'text/event-stream'
            return Response(generate_data(), mimetype='text/event-stream')
        
        # --- CHART Requirements
        @app.route("/chart-data")
        @login_required
        def get_chart_data():
            """
            Menyediakan data suhu untuk 24 jam terakhir dalam format
            yang siap digunakan oleh Chart.js.
            """
            try:
                # Tentukan rentang waktu (24 jam dari sekarang)
                conn = self.db_manager.get_connection()
                c = conn.cursor()
                
                since_time = datetime.datetime.now() - datetime.timedelta(hours=24)
                since_time_str = since_time.strftime("%Y-%m-%d %H:%M:%S")

                # Ambil data dari database
                c.execute("""
                    SELECT strftime('%H:%M', waktu), dryer_id, suhu 
                    FROM suhu 
                    WHERE waktu >= ? 
                    ORDER BY waktu ASC
                """, (since_time_str,))
                rows = c.fetchall()
                conn.close()

                if not rows:
                    return jsonify({"labels": [], "datasets": []})

                # Proses pivot data
                labels = []
                data_points = {} # Format: { "14:30": {"dryer1": 150.1, "dryer2": 152.3}, ... }

                for row in rows:
                    waktu, dryer_id, suhu = row
                    if waktu not in data_points:
                        data_points[waktu] = { "dryer1": None, "dryer2": None, "dryer3": None }
                        labels.append(waktu)
                    
                    if dryer_id in data_points[waktu]:
                        data_points[waktu][dryer_id] = suhu

                # Siapkan dataset untuk Chart.js
                dryer1_data = [data_points[t].get('dryer1') for t in labels]
                dryer2_data = [data_points[t].get('dryer2') for t in labels]
                dryer3_data = [data_points[t].get('dryer3') for t in labels]

                chart_data = {
                    "labels": labels,
                    "datasets": [
                        {
                            "label": "Dryer 1",
                            "data": dryer1_data,
                            "borderColor": "rgba(255, 99, 132, 1)",
                            "backgroundColor": "rgba(255, 99, 132, 0.2)",
                            "fill": True,
                            "tension": 0.4 # Membuat garis lebih melengkung
                        },
                        {
                            "label": "Dryer 2",
                            "data": dryer2_data,
                            "borderColor": "rgba(54, 162, 235, 1)",
                            "backgroundColor": "rgba(54, 162, 235, 0.2)",
                            "fill": True,
                            "tension": 0.4
                        },
                        {
                            "label": "Dryer 3",
                            "data": dryer3_data,
                            "borderColor": "rgba(75, 192, 192, 1)",
                            "backgroundColor": "rgba(75, 192, 192, 0.2)",
                            "fill": True,
                            "tension": 0.4
                        }
                    ]
                }
                return jsonify(chart_data)

            except Exception as e:
                logger.error(f"Error getting chart data: {e}")
                return jsonify({"error": str(e)}), 500
        
        @app.route("/data")
        def get_data_api():
            selected_date = request.args.get('date')
            # Memanggil dengan latest_only=True untuk mendapatkan 1 data terbaru
            rows = self.db_manager.get_data_by_date_pivoted(selected_date, latest_only=True)
            data = [{"waktu": r[0], "dryer1": r[1], "dryer2": r[2], "dryer3": r[3]} for r in rows]
            return jsonify(data)

        @app.route("/download")
        def download_excel():
            selected_date = request.args.get('date')
            # Memanggil tanpa parameter tambahan untuk mendapatkan semua data
            rows = self.db_manager.get_data_by_date_pivoted(selected_date)
            if not rows: return "Tidak ada data.", 404
            wb = Workbook()
            ws = wb.active
            ws.append(["Waktu (WIB)", "Dryer 1 (°C)", "Dryer 2 (°C)", "Dryer 3 (°C)"])
            for row in rows: ws.append(list(row))
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            filename = f"laporan_{selected_date}.xlsx"
            return Response(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': f'attachment;filename={filename}'})
        
        @app.route("/keepalive")
        def keepalive():
            return {"status": "alive", "timestamp": self.config.format_indonesia_time()}
        

        @app.route("/test-telegram")
        def test_telegram():
            message = f"🧪 **Test Message**\n🕐 {self.config.format_indonesia_time()}"
            self.telegram_service.send_message(message)
            return {"status": "success", "message": "Test message queued"}
        
        return app
    
    def run(self):
        # --- LOGIN SYSTEM: Membuat user awal dari secrets saat aplikasi start ---
        admin_user = os.getenv('ADMIN_USER')
        admin_pass = os.getenv('ADMIN_PASSWORD')
        if admin_user and admin_pass:
            self.db_manager.create_initial_user(admin_user, admin_pass)
        else:
            logger.warning("ADMIN_USER dan ADMIN_PASSWORD tidak diatur. Tidak dapat membuat/memverifikasi user admin.")
        # --------------------------------------------------------------------
        try:
            self.mqtt_service.connect()
            self.telegram_service.start_worker()
            self.start_background_tasks()
            app = self.create_flask_app()
            flask_thread = threading.Thread(target=lambda: app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 8080))), daemon=True)
            flask_thread.start()
            self.telegram_service.start_polling()
        except KeyboardInterrupt:
            logger.info("Shutting down...")
        finally:
            self.stop_background_tasks()
            self.telegram_service.stop_worker()
            self.mqtt_service.disconnect()

if __name__ == "__main__":
    monitor = TemperatureMonitor()
    monitor.run()